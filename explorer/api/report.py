"""
/api/report/{actor_id}  — generate and stream an Excel threat model report.

Five sheets:
  1. Actor Profile   — identity, targeting, aliases
  2. Techniques      — TTPs with STIX procedure citations
  3. NIST Controls   — per-technique control mapping with posture
  4. D3FEND          — per-technique countermeasure mapping with posture
  5. Gap Summary     — coverage stats and ranked action list

Pure sourced extraction — no LLM synthesis. Every row is traceable to
MITRE ATT&CK STIX data, the CTID NIST 800-53 crosswalk, or the D3FEND ontology.
"""

import io
from datetime import date

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from sqlalchemy.orm import Session

from core.db import get_session
from core.models import (
    Actor, Alias, ActorTechnique, ActorSoftware,
    Control, ControlPosture, D3FendPosture, D3FendTechnique,
    Software, Targeting, Technique, TechniqueControl, TechniqueD3Fend,
)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

router = APIRouter(tags=["report"])

TACTIC_ORDER = [
    "reconnaissance", "resource-development", "initial-access", "execution",
    "persistence", "privilege-escalation", "defense-evasion", "credential-access",
    "discovery", "lateral-movement", "collection", "command-and-control",
    "exfiltration", "impact", "unknown",
]

# ---------------------------------------------------------------------------
# Colours
# ---------------------------------------------------------------------------
HDR_FILL   = "1E293B"   # dark slate header
HDR_FONT   = "FFFFFF"
GREEN_FILL = "D1FAE5"   # covered / implemented
RED_FILL   = "FEE2E2"   # not deployed
GREY_FILL  = "F1F5F9"   # no mapping
WARN_FILL  = "FEF3C7"   # partial / low
ACC_FILL   = "EDE9FE"   # accent purple


def _db():
    with get_session() as s:
        yield s


@router.get("/api/report/{actor_id}")
def generate_report(actor_id: int, db: Session = Depends(_db)):
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed. Run: pip install openpyxl",
        )

    actor = db.query(Actor).filter_by(id=actor_id).first()
    if not actor:
        raise HTTPException(status_code=404, detail="Actor not found")

    # ---- Fetch all data up front ----------------------------------------

    aliases = db.query(Alias).filter_by(actor_id=actor_id).all()
    targeting = db.query(Targeting).filter_by(actor_id=actor_id).all()

    actor_techniques = (
        db.query(ActorTechnique, Technique)
        .join(Technique, ActorTechnique.technique_id == Technique.id)
        .filter(ActorTechnique.actor_id == actor_id)
        .all()
    )
    technique_ids = [tech.id for _, tech in actor_techniques]
    tech_map: dict[int, Technique] = {tech.id: tech for _, tech in actor_techniques}
    proc_map: dict[int, str] = {
        tech.id: at.procedure or ""
        for at, tech in actor_techniques
    }

    # NIST
    implemented_control_ids: set[int] = {
        r.control_id
        for r in db.query(ControlPosture).filter_by(implemented=True).all()
    }
    nist_rows = (
        db.query(TechniqueControl, Control)
        .join(Control, TechniqueControl.control_id == Control.id)
        .filter(TechniqueControl.technique_id.in_(technique_ids))
        .all()
    ) if technique_ids else []
    nist_by_technique: dict[int, list] = {}
    for tc, ctrl in nist_rows:
        nist_by_technique.setdefault(tc.technique_id, []).append((tc, ctrl))

    # D3FEND
    implemented_d3fend_ids: set[int] = {
        r.d3fend_technique_id
        for r in db.query(D3FendPosture).filter_by(implemented=True).all()
    }
    d3fend_rows = (
        db.query(TechniqueD3Fend, D3FendTechnique)
        .join(D3FendTechnique, TechniqueD3Fend.d3fend_technique_id == D3FendTechnique.id)
        .filter(TechniqueD3Fend.technique_id.in_(technique_ids))
        .all()
    ) if technique_ids else []
    d3fend_by_technique: dict[int, list] = {}
    for td, dt in d3fend_rows:
        d3fend_by_technique.setdefault(td.technique_id, []).append((td, dt))

    actor_software = (
        db.query(ActorSoftware, Software)
        .join(Software, ActorSoftware.software_id == Software.id)
        .filter(ActorSoftware.actor_id == actor_id)
        .all()
    )

    # ---- Compute gap summary --------------------------------------------

    n_covered, n_not_deployed, n_no_mapping = 0, 0, 0
    nist_covered = 0
    for tech_id in technique_ids:
        cms = d3fend_by_technique.get(tech_id, [])
        if not cms:
            n_no_mapping += 1
        elif any(dt.id in implemented_d3fend_ids for _, dt in cms):
            n_covered += 1
        else:
            n_not_deployed += 1

        ctrls = nist_by_technique.get(tech_id, [])
        if any(ctrl.id in implemented_control_ids for _, ctrl in ctrls):
            nist_covered += 1

    mappable = n_covered + n_not_deployed
    d3fend_pct = round(n_covered / mappable * 100) if mappable else 0
    nist_pct = round(nist_covered / len(technique_ids) * 100) if technique_ids else 0

    # Ranked recommendations (not_deployed techniques only)
    cm_to_open_techs: dict[int, list] = {}
    for tech_id in technique_ids:
        cms = d3fend_by_technique.get(tech_id, [])
        if not cms:
            continue
        if any(dt.id in implemented_d3fend_ids for _, dt in cms):
            continue  # already covered
        for _, dt in cms:
            if dt.id not in implemented_d3fend_ids:
                cm_to_open_techs.setdefault(dt.id, []).append(tech_id)

    recs: list[tuple] = []
    for cm_id, open_tech_ids in cm_to_open_techs.items():
        dt_rows = [dt for _, dt in d3fend_rows if dt.id == cm_id]
        if not dt_rows:
            continue
        dt = dt_rows[0]
        recs.append((len(open_tech_ids), dt.d3fend_id, dt.name, dt.tactic or "Unknown"))
    recs.sort(key=lambda r: -r[0])

    # ---- Build workbook --------------------------------------------------

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    _sheet_profile(wb, actor, aliases, targeting, actor_software, date.today())
    _sheet_techniques(wb, actor_techniques, proc_map, d3fend_by_technique, implemented_d3fend_ids, nist_by_technique, implemented_control_ids)
    _sheet_nist(wb, actor_techniques, nist_by_technique, implemented_control_ids)
    _sheet_d3fend(wb, actor_techniques, d3fend_by_technique, implemented_d3fend_ids)
    _sheet_gap_summary(wb, actor, n_covered, n_not_deployed, n_no_mapping, d3fend_pct, nist_covered, len(technique_ids), nist_pct, recs)

    # ---- Stream as file download ----------------------------------------

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = actor.name.replace(" ", "_").replace("/", "-")
    filename = f"threat_model_{safe_name}_{date.today()}.xlsx"

    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _hdr_style(cell, bg=HDR_FILL, fg=HDR_FONT):
    cell.font = Font(bold=True, color=fg, name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)


def _header_row(ws, cols: list[str], row: int = 1, bg=HDR_FILL):
    for ci, label in enumerate(cols, 1):
        cell = ws.cell(row=row, column=ci, value=label)
        _hdr_style(cell, bg=bg)
    ws.row_dimensions[row].height = 18


def _set_col_widths(ws, widths: list[int]):
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def _fill_cell(cell, fill_hex: str):
    cell.fill = PatternFill("solid", start_color=fill_hex)


def _data_style(cell, wrap=False):
    cell.font = Font(name="Arial", size=9)
    cell.alignment = Alignment(wrap_text=wrap, vertical="top")


# ---- Sheet 1: Actor Profile ----

def _sheet_profile(wb, actor, aliases, targeting, actor_software, today):
    ws = wb.create_sheet("Actor Profile")

    def kv(label, value, row):
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(bold=True, name="Arial", size=9)
        lc.fill = PatternFill("solid", start_color="E2E8F0")
        vc = ws.cell(row=row, column=2, value=value or "—")
        vc.font = Font(name="Arial", size=9)
        vc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 15

    # Title
    title = ws.cell(row=1, column=1, value=f"Threat Model Report — {actor.name}")
    title.font = Font(bold=True, name="Arial", size=14, color=HDR_FILL)
    ws.cell(row=2, column=1, value=f"Generated: {today}").font = Font(name="Arial", size=9, color="64748B")

    r = 4
    kv("Name", actor.name, r); r += 1
    kv("ATT&CK Group ID", actor.attack_group_id, r); r += 1
    kv("Country Code", actor.country_code, r); r += 1
    kv("First Seen", actor.first_seen, r); r += 1
    kv("Last Seen", actor.last_seen, r); r += 1
    kv("STIX ID", actor.stix_id, r); r += 1
    kv("MITRE URL",
       f"https://attack.mitre.org/groups/{actor.attack_group_id}/" if actor.attack_group_id else None,
       r); r += 1

    # Description
    kv("Description", actor.description, r)
    ws.row_dimensions[r].height = 80; r += 1

    # Targeting
    industries = sorted({t.value for t in targeting if t.target_type == "industry"})
    regions = sorted({t.value for t in targeting if t.target_type in ("region", "country")})
    kv("Target Industries", ", ".join(industries) or "—", r); r += 1
    kv("Target Regions", ", ".join(regions) or "—", r); r += 1

    # Aliases
    alias_parts = [
        f"{a.alias} ({a.source})" for a in sorted(aliases, key=lambda x: x.source)
        if a.alias != actor.name
    ]
    kv("Known Aliases", "\n".join(alias_parts) or "—", r)
    ws.row_dimensions[r].height = max(15, len(alias_parts) * 14); r += 1

    # Software
    tools = [sw.name for _, sw in actor_software if sw.software_type == "tool"]
    malware = [sw.name for _, sw in actor_software if sw.software_type == "malware"]
    kv("Tools", ", ".join(sorted(tools)) or "—", r); r += 1
    kv("Malware", ", ".join(sorted(malware)) or "—", r); r += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80


# ---- Sheet 2: Techniques ----

def _sheet_techniques(wb, actor_techniques, proc_map, d3fend_by_technique, implemented_d3fend_ids, nist_by_technique, implemented_control_ids):
    ws = wb.create_sheet("Techniques")
    cols = [
        "ATT&CK ID", "Name", "Tactic", "Sub-technique",
        "Technique Description", "Procedure (STIX Citation)",
        "D3FEND Status", "NIST Status",
    ]
    _header_row(ws, cols)
    ws.freeze_panes = "A2"

    sorted_at = sorted(
        actor_techniques,
        key=lambda x: (
            TACTIC_ORDER.index(x[1].tactic.split(",")[0].strip())
            if x[1].tactic and x[1].tactic.split(",")[0].strip() in TACTIC_ORDER else 99,
            x[1].attack_id,
        )
    )

    for ri, (at, tech) in enumerate(sorted_at, 2):
        cms = d3fend_by_technique.get(tech.id, [])
        if not cms:
            d3fend_status = "No D3FEND Mapping"
            d3fend_fill = GREY_FILL
        elif any(dt.id in implemented_d3fend_ids for _, dt in cms):
            d3fend_status = "Covered"
            d3fend_fill = GREEN_FILL
        else:
            d3fend_status = "Not Deployed"
            d3fend_fill = RED_FILL

        nist_ctrls = nist_by_technique.get(tech.id, [])
        if not nist_ctrls:
            nist_status = "No NIST Mapping"
            nist_fill = GREY_FILL
        elif any(ctrl.id in implemented_control_ids for _, ctrl in nist_ctrls):
            nist_status = "Covered"
            nist_fill = GREEN_FILL
        else:
            nist_status = "Not Covered"
            nist_fill = RED_FILL

        tactic = tech.tactic.split(",")[0].strip() if tech.tactic else "unknown"
        values = [
            tech.attack_id, tech.name, tactic,
            "Yes" if tech.is_subtechnique else "No",
            tech.description or "",
            proc_map.get(tech.id, ""),
            d3fend_status, nist_status,
        ]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            _data_style(cell, wrap=(ci in (5, 6)))
        _fill_cell(ws.cell(row=ri, column=7), d3fend_fill)
        _fill_cell(ws.cell(row=ri, column=8), nist_fill)
        ws.row_dimensions[ri].height = 60 if proc_map.get(tech.id) else 30

    _set_col_widths(ws, [12, 30, 22, 14, 50, 60, 18, 18])


# ---- Sheet 3: NIST Controls ----

def _sheet_nist(wb, actor_techniques, nist_by_technique, implemented_control_ids):
    ws = wb.create_sheet("NIST Controls")
    cols = ["ATT&CK ID", "Technique Name", "Control ID", "Control Name", "Family", "Mapping Type", "Implemented"]
    _header_row(ws, cols)
    ws.freeze_panes = "A2"

    ri = 2
    for at, tech in sorted(actor_techniques, key=lambda x: x[1].attack_id):
        ctrls = nist_by_technique.get(tech.id, [])
        if not ctrls:
            for ci, val in enumerate([tech.attack_id, tech.name, "—", "No NIST mapping", "—", "—", "—"], 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                _data_style(cell)
                _fill_cell(cell, GREY_FILL)
            ri += 1
            continue
        for tc, ctrl in sorted(ctrls, key=lambda x: x[1].control_id):
            impl = ctrl.id in implemented_control_ids
            values = [
                tech.attack_id, tech.name,
                ctrl.control_id, ctrl.name, ctrl.control_group or "—",
                tc.mapping_type, "Yes" if impl else "No",
            ]
            for ci, val in enumerate(values, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                _data_style(cell)
            _fill_cell(ws.cell(row=ri, column=7), GREEN_FILL if impl else RED_FILL)
            ri += 1

    _set_col_widths(ws, [12, 30, 12, 40, 10, 16, 12])


# ---- Sheet 4: D3FEND ----

def _sheet_d3fend(wb, actor_techniques, d3fend_by_technique, implemented_d3fend_ids):
    ws = wb.create_sheet("D3FEND")
    cols = ["ATT&CK ID", "Technique Name", "D3FEND ID", "Countermeasure Name", "D3FEND Tactic", "Deployed", "Coverage Bucket"]
    _header_row(ws, cols)
    ws.freeze_panes = "A2"

    ri = 2
    for at, tech in sorted(actor_techniques, key=lambda x: x[1].attack_id):
        cms = d3fend_by_technique.get(tech.id, [])
        if not cms:
            for ci, val in enumerate([tech.attack_id, tech.name, "—", "No D3FEND mapping", "—", "—", "no_mapping"], 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                _data_style(cell)
                _fill_cell(cell, GREY_FILL)
            ri += 1
            continue

        any_impl = any(dt.id in implemented_d3fend_ids for _, dt in cms)
        for _, dt in sorted(cms, key=lambda x: x[1].d3fend_id):
            deployed = dt.id in implemented_d3fend_ids
            bucket = "covered" if any_impl else "not_deployed"
            values = [
                tech.attack_id, tech.name,
                dt.d3fend_id, dt.name, dt.tactic or "Unknown",
                "Yes" if deployed else "No", bucket,
            ]
            for ci, val in enumerate(values, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                _data_style(cell)
            _fill_cell(ws.cell(row=ri, column=6), GREEN_FILL if deployed else RED_FILL)
            _fill_cell(ws.cell(row=ri, column=7),
                       GREEN_FILL if bucket == "covered" else RED_FILL)
            ri += 1

    _set_col_widths(ws, [12, 30, 12, 36, 14, 10, 16])


# ---- Sheet 5: Gap Summary ----

def _sheet_gap_summary(wb, actor, n_covered, n_not_deployed, n_no_mapping, d3fend_pct, nist_covered, total_techs, nist_pct, recs):
    ws = wb.create_sheet("Gap Summary")

    def section_hdr(row, label):
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = Font(bold=True, name="Arial", size=10, color=HDR_FILL)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 20

    def kv(row, label, value, fill=None):
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(bold=True, name="Arial", size=9)
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = Font(name="Arial", size=9)
        if fill:
            _fill_cell(vc, fill)

    r = 1
    title = ws.cell(row=r, column=1, value=f"Gap Summary — {actor.name}")
    title.font = Font(bold=True, name="Arial", size=13, color=HDR_FILL)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 2

    section_hdr(r, "D3FEND Detection Coverage"); r += 1
    mappable = n_covered + n_not_deployed
    kv(r, "Coverage %", f"{d3fend_pct}%",
       fill=GREEN_FILL if d3fend_pct >= 70 else (WARN_FILL if d3fend_pct >= 40 else RED_FILL)); r += 1
    kv(r, "Covered (deployed countermeasure)", n_covered, fill=GREEN_FILL); r += 1
    kv(r, "Not Deployed (countermeasure exists)", n_not_deployed, fill=RED_FILL); r += 1
    kv(r, "No D3FEND Mapping", n_no_mapping, fill=GREY_FILL); r += 1
    kv(r, "Total techniques", n_covered + n_not_deployed + n_no_mapping); r += 1
    kv(r, "Denominator (mappable only)", mappable); r += 2

    section_hdr(r, "NIST 800-53 Compliance Coverage"); r += 1
    kv(r, "Coverage %", f"{nist_pct}%",
       fill=GREEN_FILL if nist_pct >= 70 else (WARN_FILL if nist_pct >= 40 else RED_FILL)); r += 1
    kv(r, "Covered (implemented control)", nist_covered, fill=GREEN_FILL); r += 1
    kv(r, "Not Covered", total_techs - nist_covered, fill=RED_FILL); r += 1
    kv(r, "Total techniques", total_techs); r += 2

    section_hdr(r, "Ranked Action List (top countermeasures by gap closure)"); r += 1
    _header_row(ws, ["Rank", "D3FEND ID", "Countermeasure", "Tactic", "Techniques Closed"], row=r, bg="334155")
    r += 1
    for i, (count, d3id, name, tactic) in enumerate(recs[:20], 1):
        ws.cell(row=r, column=1, value=i).font = Font(name="Arial", size=9)
        ws.cell(row=r, column=2, value=d3id).font = Font(name="Arial", size=9)
        ws.cell(row=r, column=3, value=name).font = Font(name="Arial", size=9)
        ws.cell(row=r, column=4, value=tactic).font = Font(name="Arial", size=9)
        cnt_cell = ws.cell(row=r, column=5, value=count)
        cnt_cell.font = Font(bold=True, name="Arial", size=9)
        _fill_cell(cnt_cell, ACC_FILL)
        r += 1

    _set_col_widths(ws, [34, 14, 40, 14, 18])
